#!/usr/bin/env python3
"""
Compute Standardised Claim Ratios (SCR) for pain-related prescribing
using indirect age-sex standardisation.

Method (following Taira et al, Lancet Reg Health West Pac 2021):
  1. Compute national age-sex-specific prescription rates (NDB sex/age table)
  2. Apply rates to each prefecture's age-sex population structure (e-Stat)
  3. SCR = (Observed / Expected) × 100

Data sources:
  - NDB Open Data 10th ed: outpatient_drugs_sexage.xlsx (national, by sex × age)
  - NDB Open Data 10th ed: outpatient_drugs_prefecture.xlsx (by prefecture)
  - NDB Open Data 10th ed: inpatient_drugs_sexage.xlsx (national, by sex × age)
  - NDB Open Data 10th ed: inpatient_drugs_prefecture.xlsx (by prefecture)
  - e-Stat: population_pref_age_sex_2023.xlsx (2023 Oct population estimates)
"""

import os
import json
import csv
import numpy as np
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
DATA_DIR = os.path.join(BASE_DIR, 'data')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ============================================================
# AGE GROUP DEFINITIONS
# ============================================================
# NDB uses 21 age groups per sex: 0-4, 5-9, ..., 95-99, 100+
# Population data uses 18 groups: 0-4, 5-9, ..., 80-84, 85+
# We harmonise to 18 groups by merging NDB 85-89, 90-94, 95-99, 100+ into 85+

NDB_AGE_LABELS = [
    '0～4歳', '5～9歳', '10～14歳', '15～19歳', '20～24歳',
    '25～29歳', '30～34歳', '35～39歳', '40～44歳', '45～49歳',
    '50～54歳', '55～59歳', '60～64歳', '65～69歳', '70～74歳',
    '75～79歳', '80～84歳', '85～89歳', '90～94歳', '95～99歳', '100歳以上'
]

# Mapping NDB 21 groups -> 18 harmonised groups (merge indices 17-20 into 17)
N_HARMONISED = 18  # 0-4 through 80-84 = 17 groups + 85+ = 18 groups

def harmonise_ndb_to_18(vals_21):
    """Merge NDB 21 age groups into 18 by combining 85+ groups."""
    out = list(vals_21[:17])  # 0-4 through 80-84
    out.append(sum(vals_21[17:21]))  # 85+ = 85-89 + 90-94 + 95-99 + 100+
    return out


def safe_float(val):
    if val is None or str(val).strip() in ('-', '－', '‐', ''):
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


# ============================================================
# 1. LOAD POPULATION DATA (prefecture × age × sex)
# ============================================================
print("=" * 70)
print("1. Loading prefecture population by age and sex (2023)")
print("=" * 70)

wb_pop = openpyxl.load_workbook(
    os.path.join(DATA_DIR, 'population_pref_age_sex_2023.xlsx'), read_only=True)
sheet_pop = wb_pop['第10表']

# Structure: rows from index 7 onward
# Col 9 = pop type (総人口), Col 10 = pref code (00000=national, 01000-47000),
# Col 12 = sex (男女計, 男, 女), Cols 15=total, 16-33=age groups (18 groups)
# Units: 千人 (thousands)

pop_data = {}  # {pref_code: {'male': [18 ages], 'female': [18 ages], 'total': [18 ages]}}
national_pop = {}  # {'male': [18 ages], 'female': [18 ages], 'total': [18 ages]}

sex_map = {'男女計': 'total', '男': 'male', '女': 'female'}

for i, row in enumerate(sheet_pop.iter_rows(values_only=True)):
    vals = list(row)
    if i < 7:
        continue
    pop_type = str(vals[9]).strip() if vals[9] else ''
    if pop_type != '総人口':
        continue
    pref_str = str(vals[10]).strip() if vals[10] else ''
    if not pref_str or not pref_str.isdigit():
        continue
    pref_code = int(pref_str) // 1000  # e.g., 01000 -> 1, 00000 -> 0
    sex_label = str(vals[12]).strip() if vals[12] else ''
    sex_key = sex_map.get(sex_label)
    if sex_key is None:
        continue

    # Age groups in cols 16-33 (18 groups), in thousands
    age_vals = [safe_float(vals[j]) * 1000 for j in range(16, 34)]  # Convert thousands to persons

    if pref_code == 0:
        national_pop[sex_key] = age_vals
    else:
        if pref_code not in pop_data:
            pop_data[pref_code] = {}
        pop_data[pref_code][sex_key] = age_vals

wb_pop.close()
print(f"  Loaded {len(pop_data)} prefectures")
print(f"  National total: {sum(national_pop['total']):,.0f}")
print(f"  Age groups: {N_HARMONISED}")


# ============================================================
# 2. LOAD NDB SEX-AGE DRUG DATA (national level)
# ============================================================
print("\n" + "=" * 70)
print("2. Loading NDB sex-age drug data (outpatient + inpatient)")
print("=" * 70)


def extract_drug_sexage(filepath, sheet_name, keywords, label):
    """Extract national drug prescriptions by sex × age group.

    Returns: {'male': [18 ages], 'female': [18 ages], 'total': [18 ages]}
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet = wb[sheet_name]

    male_total = [0.0] * 21
    female_total = [0.0] * 21
    count = 0

    for row in sheet.iter_rows(values_only=True):
        vals = list(row)
        if len(vals) < 51:
            continue
        drug_name = str(vals[3]) if vals[3] else ''
        if not any(kw in drug_name for kw in keywords):
            continue
        count += 1
        # Male: cols 9-29 (21 age groups)
        for j in range(21):
            male_total[j] += safe_float(vals[9 + j])
        # Female: cols 30-50 (21 age groups)
        for j in range(21):
            female_total[j] += safe_float(vals[30 + j])

    wb.close()

    # Harmonise to 18 groups
    male_18 = harmonise_ndb_to_18(male_total)
    female_18 = harmonise_ndb_to_18(female_total)
    total_18 = [m + f for m, f in zip(male_18, female_18)]

    national_total = sum(total_18)
    print(f"  {label}: {count} entries, national total: {national_total:,.0f}")
    return {'male': male_18, 'female': female_18, 'total': total_18}


def extract_drug_prefecture(filepath, sheet_name, keywords, label):
    """Extract drug prescriptions by prefecture. Returns: {pref_code: quantity}"""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet = wb[sheet_name]

    pref_totals = {i: 0.0 for i in range(1, 48)}
    national = 0.0
    count = 0
    col_start = 9  # 0-indexed column for prefecture 01

    for row in sheet.iter_rows(values_only=True):
        vals = list(row)
        if len(vals) < 56:
            continue
        drug_name = str(vals[3]) if vals[3] else ''
        if not any(kw in drug_name for kw in keywords):
            continue
        count += 1
        national += safe_float(vals[8])
        for pc in range(1, 48):
            col = col_start + (pc - 1)
            if col < len(vals):
                pref_totals[pc] += safe_float(vals[col])

    wb.close()
    print(f"  {label} (pref): {count} entries, national total: {national:,.0f}")
    return pref_totals


# --- Drug categories ---
# Phase 1: Acute perioperative analgesics (inpatient)
ANALGESIC_KEYWORDS_INPATIENT = {
    'nsaids': ['ロキソプロフェン', 'ロキソニン', 'セレコキシブ', 'セレコックス',
               'ジクロフェナク', 'ボルタレン', 'アセトアミノフェン', 'カロナール',
               'イブプロフェン', 'ナプロキセン', 'メロキシカム', 'モービック',
               'インドメタシン', 'フルルビプロフェン', 'エトドラク', 'ハイペン',
               'ピロキシカム', 'フェルデン', 'ザルトプロフェン', 'ペオン',
               'オキサプロジン', 'アルミノプロフェン'],
    'opioids': ['オキシコドン', 'オキシコンチン', 'オキノーム',
                'フェンタニル', 'デュロテップ', 'アブストラル', 'イーフェン',
                'モルヒネ', 'MSコンチン', 'オプソ', 'カディアン',
                'コデイン', 'トラマドール', 'トラマール', 'ワントラム',
                'タペンタドール', 'ペンタゾシン', 'ブプレノルフィン',
                'ヒドロモルフォン', 'ナルサス', 'ナルラピド',
                'メサドン'],
}

# Phase 2: Neuropathic pain drugs (outpatient)
NEUROPATHIC_KEYWORDS = ['プレガバリン', 'リリカ', 'ミロガバリン', 'タリージェ',
                         'デュロキセチン', 'サインバルタ',
                         'トラマドール', 'トラムセット',
                         'ノイロトロピン', 'ワクシニア']

# Confounder proxies
DIABETES_KEYWORDS = ['メトホルミン', 'グリメピリド', 'アマリール', 'シタグリプチン',
                     'ジャヌビア', 'リナグリプチン', 'トラゼンタ', 'テネリグリプチン',
                     'テネリア', 'アログリプチン', 'ネシーナ', 'ビルダグリプチン',
                     'エクア', 'サキサグリプチン', 'オングリザ',
                     'エンパグリフロジン', 'ジャディアンス', 'ダパグリフロジン',
                     'フォシーガ', 'カナグリフロジン', 'カナグル',
                     'イプラグリフロジン', 'スーグラ', 'ルセオグリフロジン',
                     'ピオグリタゾン', 'アクトス', 'ボグリボース', 'ベイスン',
                     'ミグリトール', 'セイブル',
                     'グリクラジド', 'グリベンクラミド', 'ダオニール', 'オイグルコン',
                     'レパグリニド', 'シュアポスト', 'ナテグリニド', 'スターシス',
                     'ファスティック', 'ミチグリニド', 'グルファスト', 'メトグルコ',
                     'グルベス', 'マリゼブ', 'オマリグリプチン', 'ザファテック',
                     'トレラグリプチン']

outpatient_sexage_file = os.path.join(DATA_DIR, 'outpatient_drugs_sexage.xlsx')
outpatient_pref_file = os.path.join(DATA_DIR, 'outpatient_drugs_prefecture.xlsx')
inpatient_sexage_file = os.path.join(DATA_DIR, 'inpatient_drugs_sexage.xlsx')
inpatient_pref_file = os.path.join(DATA_DIR, 'inpatient_drugs_prefecture.xlsx')
outpatient_sheet = '内服薬 外来 (院外)'
inpatient_sheet = '内服薬 入院'

# --- Extract sex-age data (national) ---
print("\n[A] Outpatient neuropathic pain drugs (sex-age, national)")
neuro_sexage = extract_drug_sexage(outpatient_sexage_file, outpatient_sheet,
                                    NEUROPATHIC_KEYWORDS, 'Neuropathic pain')

print("\n[B] Outpatient diabetes drugs (sex-age, national)")
diabetes_sexage = extract_drug_sexage(outpatient_sexage_file, outpatient_sheet,
                                       DIABETES_KEYWORDS, 'Diabetes drugs')

print("\n[C] Inpatient analgesics (sex-age, national)")
all_analgesic_kw = ANALGESIC_KEYWORDS_INPATIENT['nsaids'] + ANALGESIC_KEYWORDS_INPATIENT['opioids']
analgesic_sexage = extract_drug_sexage(inpatient_sexage_file, inpatient_sheet,
                                        all_analgesic_kw, 'Analgesics (inpatient)')

# --- Extract prefecture data (observed) ---
print("\n[D] Outpatient neuropathic pain drugs (by prefecture)")
neuro_pref = extract_drug_prefecture(outpatient_pref_file, outpatient_sheet,
                                      NEUROPATHIC_KEYWORDS, 'Neuropathic pain')

print("\n[E] Outpatient diabetes drugs (by prefecture)")
diabetes_pref = extract_drug_prefecture(outpatient_pref_file, outpatient_sheet,
                                         DIABETES_KEYWORDS, 'Diabetes drugs')

print("\n[F] Inpatient analgesics (by prefecture)")
analgesic_pref = extract_drug_prefecture(inpatient_pref_file, inpatient_sheet,
                                          all_analgesic_kw, 'Analgesics (inpatient)')


# ============================================================
# 3. COMPUTE SCR (indirect standardisation)
# ============================================================
print("\n" + "=" * 70)
print("3. Computing Standardised Claim Ratios (SCR)")
print("=" * 70)


def compute_scr(drug_sexage, drug_pref, pop_national, pop_pref_data, label):
    """
    SCR = (Observed / Expected) × 100

    Expected_i = sum_a,s [ (national_rate_a,s) × (pop_i_a,s) ]
    where national_rate_a,s = national_drug_a,s / national_pop_a,s
    """
    # Compute national age-sex-specific rates
    rates = {'male': [0.0] * N_HARMONISED, 'female': [0.0] * N_HARMONISED}
    for sex in ('male', 'female'):
        for a in range(N_HARMONISED):
            pop = pop_national[sex][a]
            prescriptions = drug_sexage[sex][a]
            rates[sex][a] = prescriptions / pop if pop > 0 else 0.0

    results = {}
    for pc in range(1, 48):
        if pc not in pop_pref_data:
            continue
        observed = drug_pref[pc]

        # Expected = sum of (national rate × prefecture population) for each age-sex group
        expected = 0.0
        for sex in ('male', 'female'):
            if sex not in pop_pref_data[pc]:
                continue
            for a in range(N_HARMONISED):
                expected += rates[sex][a] * pop_pref_data[pc][sex][a]

        scr = (observed / expected) * 100 if expected > 0 else np.nan
        results[pc] = {
            'observed': observed,
            'expected': expected,
            'scr': scr,
        }

    # Summary
    scr_vals = [r['scr'] for r in results.values() if not np.isnan(r['scr'])]
    print(f"\n  {label}:")
    print(f"    Prefectures: {len(scr_vals)}")
    print(f"    SCR range: {min(scr_vals):.1f} – {max(scr_vals):.1f}")
    print(f"    SCR mean: {np.mean(scr_vals):.1f}, median: {np.median(scr_vals):.1f}")
    return results


neuro_scr = compute_scr(neuro_sexage, neuro_pref, national_pop, pop_data,
                         'Neuropathic pain drugs (outpatient)')
diabetes_scr = compute_scr(diabetes_sexage, diabetes_pref, national_pop, pop_data,
                            'Diabetes drugs (outpatient)')
analgesic_scr = compute_scr(analgesic_sexage, analgesic_pref, national_pop, pop_data,
                             'Analgesics (inpatient)')


# ============================================================
# 4. SAVE RESULTS
# ============================================================
print("\n" + "=" * 70)
print("4. Saving results")
print("=" * 70)

PREF_NAMES = {
    1:'北海道',2:'青森県',3:'岩手県',4:'宮城県',5:'秋田県',6:'山形県',7:'福島県',
    8:'茨城県',9:'栃木県',10:'群馬県',11:'埼玉県',12:'千葉県',13:'東京都',14:'神奈川県',
    15:'新潟県',16:'富山県',17:'石川県',18:'福井県',19:'山梨県',20:'長野県',
    21:'岐阜県',22:'静岡県',23:'愛知県',24:'三重県',25:'滋賀県',26:'京都府',
    27:'大阪府',28:'兵庫県',29:'奈良県',30:'和歌山県',31:'鳥取県',32:'島根県',
    33:'岡山県',34:'広島県',35:'山口県',36:'徳島県',37:'香川県',38:'愛媛県',
    39:'高知県',40:'福岡県',41:'佐賀県',42:'長崎県',43:'熊本県',44:'大分県',
    45:'宮崎県',46:'鹿児島県',47:'沖縄県'
}

REGION_MAP = {
    1:'北海道', 2:'東北',3:'東北',4:'東北',5:'東北',6:'東北',7:'東北',
    8:'関東',9:'関東',10:'関東',11:'関東',12:'関東',13:'関東',14:'関東',
    15:'北陸・甲信越',16:'北陸・甲信越',17:'北陸・甲信越',18:'北陸・甲信越',
    19:'北陸・甲信越',20:'北陸・甲信越',
    21:'東海',22:'東海',23:'東海',24:'東海',
    25:'近畿',26:'近畿',27:'近畿',28:'近畿',29:'近畿',30:'近畿',
    31:'中国',32:'中国',33:'中国',34:'中国',35:'中国',
    36:'四国',37:'四国',38:'四国',39:'四国',
    40:'九州・沖縄',41:'九州・沖縄',42:'九州・沖縄',43:'九州・沖縄',
    44:'九州・沖縄',45:'九州・沖縄',46:'九州・沖縄',47:'九州・沖縄'
}

rows = []
for pc in range(1, 48):
    row = {
        'pref_code': pc,
        'pref_name': PREF_NAMES[pc],
        'region': REGION_MAP[pc],
        'is_tohoku': 1 if 2 <= pc <= 7 else 0,
        'analgesic_observed': analgesic_scr[pc]['observed'],
        'analgesic_expected': analgesic_scr[pc]['expected'],
        'analgesic_scr': analgesic_scr[pc]['scr'],
        'neuropathic_observed': neuro_scr[pc]['observed'],
        'neuropathic_expected': neuro_scr[pc]['expected'],
        'neuropathic_scr': neuro_scr[pc]['scr'],
        'diabetes_observed': diabetes_scr[pc]['observed'],
        'diabetes_expected': diabetes_scr[pc]['expected'],
        'diabetes_scr': diabetes_scr[pc]['scr'],
    }
    rows.append(row)

# CSV
csv_path = os.path.join(OUTPUT_DIR, 'scr_results.csv')
with open(csv_path, 'w', encoding='utf-8', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
print(f"  CSV: {csv_path}")

# JSON summary for manuscript scripts
scr_summary = {
    'analgesic_inpatient': {
        'scr_by_prefecture': {pc: analgesic_scr[pc]['scr'] for pc in range(1, 48)},
        'scr_range': [
            min(analgesic_scr[pc]['scr'] for pc in range(1, 48)),
            max(analgesic_scr[pc]['scr'] for pc in range(1, 48)),
        ],
        'scr_mean': float(np.mean([analgesic_scr[pc]['scr'] for pc in range(1, 48)])),
        'scr_tohoku_mean': float(np.mean([analgesic_scr[pc]['scr'] for pc in range(2, 8)])),
        'scr_non_tohoku_mean': float(np.mean([
            analgesic_scr[pc]['scr'] for pc in range(1, 48) if not (2 <= pc <= 7)
        ])),
        'variation_ratio': (
            max(analgesic_scr[pc]['scr'] for pc in range(1, 48)) /
            min(analgesic_scr[pc]['scr'] for pc in range(1, 48))
        ),
    },
    'neuropathic_outpatient': {
        'scr_by_prefecture': {pc: neuro_scr[pc]['scr'] for pc in range(1, 48)},
        'scr_range': [
            min(neuro_scr[pc]['scr'] for pc in range(1, 48)),
            max(neuro_scr[pc]['scr'] for pc in range(1, 48)),
        ],
        'scr_mean': float(np.mean([neuro_scr[pc]['scr'] for pc in range(1, 48)])),
        'scr_tohoku_mean': float(np.mean([neuro_scr[pc]['scr'] for pc in range(2, 8)])),
        'scr_non_tohoku_mean': float(np.mean([
            neuro_scr[pc]['scr'] for pc in range(1, 48) if not (2 <= pc <= 7)
        ])),
        'variation_ratio': (
            max(neuro_scr[pc]['scr'] for pc in range(1, 48)) /
            min(neuro_scr[pc]['scr'] for pc in range(1, 48))
        ),
    },
    'diabetes_outpatient': {
        'scr_by_prefecture': {pc: diabetes_scr[pc]['scr'] for pc in range(1, 48)},
        'scr_range': [
            min(diabetes_scr[pc]['scr'] for pc in range(1, 48)),
            max(diabetes_scr[pc]['scr'] for pc in range(1, 48)),
        ],
        'scr_mean': float(np.mean([diabetes_scr[pc]['scr'] for pc in range(1, 48)])),
    },
    'method': 'Indirect age-sex standardisation using NDB national age-sex-specific '
              'prescription rates and e-Stat 2023 October population estimates.',
    'age_groups': N_HARMONISED,
    'population_source': 'Statistics Bureau of Japan, Population Estimates 2023',
    'ndb_edition': '10th (April 2023 – March 2024)',
}

json_path = os.path.join(OUTPUT_DIR, 'scr_summary.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(scr_summary, f, ensure_ascii=False, indent=2)
print(f"  JSON: {json_path}")

# Print key results
print("\n" + "=" * 70)
print("KEY SCR RESULTS")
print("=" * 70)
for cat, key in [('Analgesics (inpatient)', 'analgesic_inpatient'),
                  ('Neuropathic pain (outpatient)', 'neuropathic_outpatient')]:
    s = scr_summary[key]
    print(f"\n  {cat}:")
    print(f"    SCR range: {s['scr_range'][0]:.1f} – {s['scr_range'][1]:.1f}")
    print(f"    Variation ratio: {s['variation_ratio']:.2f}")
    print(f"    Tohoku mean SCR: {s['scr_tohoku_mean']:.1f}")
    print(f"    Non-Tohoku mean SCR: {s['scr_non_tohoku_mean']:.1f}")

print("\nDone.")
